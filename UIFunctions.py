import re
import os
import sys
import openpyxl
import numpy as np
import pandas as pd
import validation as valid
import helpfunctions as hf
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment
from PyQt5.QtWidgets import QApplication
from compoundNamesGUI import AssignValuesWindow
from cellNamesGUI import AssignNamesValuesWindow


def get_sheet_name(cell_name: str, only_avg: bool, control_treatment: bool, fixed_col: str):
    """
    This function generate the sheet name based on different parameters.

    :param cell_name: The name of the cell line.
    :param only_avg: Flag indicating whether only average data is considered.
    :param control_treatment: Flag indicating whether control treatment is applied.
    :param fixed_col: The name of the fixed column.
    :return: The generated sheet name.
    """
    sheet_name = cell_name
    if only_avg and control_treatment:
        return sheet_name + '_CT_AVG_by_' + fixed_col
    elif not only_avg and control_treatment:
        return sheet_name + '_CT_by_' + fixed_col
    elif only_avg and not control_treatment:
        return sheet_name + '_AVG_by_' + fixed_col
    else:
        return sheet_name + '_by_' + fixed_col


def pop_up_cell_GUI(df: pd.DataFrame):
    """
    This function create a pop-up GUI window to assign names to cell lines.

    :param df: The input dataframe.
    :return: A list of assigned names for the cell lines.
    """
    cell_line_list = df['cell_line_name'].unique().tolist()
    app_names = QApplication.instance()  # Retrieve the existing QApplication instance
    if app_names is None:
        app_names = QApplication(sys.argv)
    else:
        app_names.quit()  # Terminate the existing event loop

    name_window = AssignNamesValuesWindow(cell_line_list)
    name_window.show()
    app_names.exec_()
    cell_line_list = name_window.result
    name_window.close()
    del name_window
    return cell_line_list


def pop_up_compound_GUI(pairs_df: pd.DataFrame, cell_name: str):
    """
    This function create a pop-up GUI window to assign names to compound names.

    :param pairs_df: The input pairs dataframe.
    :param cell_name: The name of the cell line.
    :return: A list of assigned names for the compound names.
    """
    control_types = ['CONTROL', 'DMSO', 'PBS']
    compound_name = pairs_df['compound_name'].unique()

    control_list = list(set(compound_name) & set(control_types))
    inhibitor_list = list(set(compound_name) - set(control_list))
    app = QApplication.instance()  # Retrieve the existing QApplication instance
    if app is None:
        app = QApplication(sys.argv)
    else:
        app.quit()  # Terminate the existing event loop

    window = AssignValuesWindow(control_list, inhibitor_list, cell_name)
    window.show()
    app.exec_()
    compounds_list = window.result
    window.close()

    return compounds_list


def set_column_width_and_alignment(worksheet):
    """
    Set the column width and alignment in the given worksheet to fit the content.

    :param worksheet: The worksheet to set the column width.
    """
    alignment = Alignment(horizontal='center')
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = alignment
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 1) * 1.1
        worksheet.column_dimensions[column].width = adjusted_width


def parse_time_string(time_str: str) -> int:
    """
    This function parses a time string and returns the time value in minutes.

    :param time_str: The time string to be parsed.
    :return: The parsed time value in minutes.
    """
    time_value = int(re.findall(r'\d+', time_str)[0])
    if "hr" in time_str:
        time_value *= 60
    return time_value


def parse_measurement_string(measurement_str: str) -> float:
    """
    This function parses a measurement string and converts it to a standardized unit.

    :param measurement_str: The measurement string to be parsed.
    :return: The parsed measurement value in the standardized unit.
    """
    if measurement_str == '-0-':
        return 0

    units = {'nM': 1e-9, 'uM': 1e-6, 'mM': 1e-3, 'ug/ml': 1}

    value_str, unit_str = re.findall(r'([\d.-]+)(\D*)', measurement_str)[0]
    value = float(value_str)
    unit = unit_str.strip()

    return value * units.get(unit, 0)


def create_bars(processes_values: dict, sorted_pairs: dict, pairs_dict: dict, pair: list, fixed_col: str,
                cell_path: str, cell_name: str):
    """
    This function create bars for given cell

    :param processes_values: Array of all the relevant data
    :param sorted_pairs: the pairs array sorted by the fixed call values
    :param pairs_dict: The data dictionary
    :param pair: relevant pair
    :param fixed_col: The fixed column the user chose
    :param cell_path: The path of the cell
    :param cell_name: The name of the cell line
    """
    print(f"Start creating bars for {cell_name}({sorted_pairs[0][1]}, {sorted_pairs[0][2]})..")
    for process_key in processes_values.keys():
        pair1, pair2, fixed_col_values = [], [], []
        values = processes_values[process_key]
        for val in values:
            pair1.append(val[0])
            pair2.append(val[1])
        for s_pair in sorted_pairs:
            df = pairs_dict[s_pair]
            if process_key in df.columns:
                fixed_col_values.append(s_pair[3])

        # Set the width of the bars
        bar_width = 0.2
        x_indices = np.arange(len(fixed_col_values))

        plt.bar(x_indices, pair1, width=bar_width, align='center', alpha=0.5, label=pair[0])
        plt.bar(x_indices + bar_width, pair2, width=bar_width, align='center', alpha=0.5, label=pair[1])

        plt.xticks(x_indices + bar_width / 2, fixed_col_values)

        for i in range(len(fixed_col_values)):
            plt.text(x_indices[i], pair1[i] + 0.1, str(round(pair1[i], 4)), ha='center', fontsize=8)
            plt.text(x_indices[i] + bar_width, pair2[i] + 0.1, str(round(pair2[i], 4)), ha='center', fontsize=8)

        if pair1[0] < 0:
            plt.gca().invert_yaxis()
        plt.xlabel(fixed_col)
        plt.ylabel('values')
        plt.legend()
        title = f'{pair[0]} {pair[1]} - PROCESS {process_key}'
        plt.title(title)
        plt.legend(loc='lower right')
        save_directory = os.path.join(cell_path, 'Bars', sorted_pairs[0][1] + ", " + sorted_pairs[0][2])
        os.makedirs(save_directory, exist_ok=True)
        save_path = os.path.join(save_directory, f' Process {process_key} - by {fixed_col}.png')
        plt.savefig(save_path, dpi=300)
        plt.close()

        print(
            f"bar {sorted_pairs[0][1]}, {sorted_pairs[0][2]}, process {process_key} saved successfully")


def create_graphs(process_sum: dict, sorted_pairs: dict, pairs_dict: dict, pair: list, fixed_col: str, cell_path: str,
                  cell_name: str):
    """
    This function create graphs for given cell

    :param process_sum: Dictionary of all the relevant data
    :param sorted_pairs: the pairs dictionary sorted by the fixed call values
    :param pairs_dict: The data dictionary
    :param pair: the relevant pair
    :param fixed_col: The fixed column the user chose
    :param cell_path: The path of the cell
    :param cell_name: The name of the cell line
    """
    print(f"Start creating graphs for {cell_name}({sorted_pairs[0][1]}, {sorted_pairs[0][2]})..")
    for process_key in process_sum.keys():
        pair1, pair2, fixed_col_values = [], [], []
        values = process_sum[process_key]
        for val in values:
            pair1.append(val[0])
            pair2.append(val[1])
        # find the relevant fixed_col_values for the values
        for s_pair in sorted_pairs:
            df = pairs_dict[s_pair]
            if process_key in df.columns:
                fixed_col_values.append(s_pair[3])

        fig, ax = plt.subplots()
        ax.plot(fixed_col_values, pair1, marker='o', label=pair[0])
        ax.plot(fixed_col_values, pair2, marker='o', label=pair[1])
        for i, txt in enumerate(range(len(fixed_col_values))):
            plt.text(fixed_col_values[i], pair1[i] + 0.04, str(round(pair1[i], 4)), fontsize=8)
            plt.text(fixed_col_values[i], pair2[i] + 0.04, str(round(pair2[i], 4)), fontsize=8)

        ax.set_xlabel(fixed_col)
        ax.set_ylabel('values')
        ax.legend()
        title = f'{pair[0]} {pair[1]} - PROCESS {process_key}'
        ax.set_title(title)

        data_min = min(min(pair1), min(pair2))
        data_max = max(max(pair1), max(pair2))
        plt.ylim(data_min - 0.5, data_max + 0.5)

        save_directory = os.path.join(cell_path, 'Graphs', sorted_pairs[0][1] + ", " + sorted_pairs[0][2])
        os.makedirs(save_directory, exist_ok=True)
        save_path = os.path.join(save_directory, f' Process {process_key} - by {fixed_col}.png')
        plt.savefig(save_path, dpi=300)
        plt.close(fig)

        print(f"graph {sorted_pairs[0][1]}, {sorted_pairs[0][2]}, process {process_key} saved successfully")


def create_plots(pairs_dict: dict, cell_name: str, fixed_col: str, cell_path: str):
    """
    The function create graphs for pairs from the data, where there is at least 2 values for process

    :param pairs_dict: The data dictionary
    :param cell_name: The name of the cell line
    :param fixed_col: The fixed column the user chose
    :param cell_path: The path of the cell
    """
    graphs = False
    pairs = list(set((key[1], key[2]) for key in pairs_dict if key[1] != key[2]))
    for pair in pairs:
        # takes all the matching pairs from the dictionary
        matching_pairs = [key for key in pairs_dict if key[1] == pair[0] and key[2] == pair[1]]
        if matching_pairs and len(matching_pairs) > 1:
            if fixed_col == 'time':
                sorted_pairs = sorted(matching_pairs, key=lambda key: parse_time_string(key[3]))  # sorted by time
            else:
                sorted_pairs = sorted(matching_pairs, key=lambda key: parse_measurement_string(key[3]))
            process_sum, processes_values = {}, {}
            # iterate over the keys and extract the df and their values pair process
            for key in sorted_pairs:
                df = pairs_dict[key]

                for process in df.columns[3:]:
                    if process not in processes_values.keys():
                        processes_values[process] = [df[process].tolist()[:2]]
                    else:
                        processes_values[process].extend([df[process].tolist()[:2]])

            create_bars(processes_values, sorted_pairs, pairs_dict, pair, fixed_col, cell_path, cell_name)

            for process_key, value in processes_values.items():
                if len(value) > 1:
                    process_sum[process_key] = value

            # If process_sum has at least 2 values for the process, create a graph
            if process_sum:
                graphs = True
                print(f"Start creating graphs for {cell_name}({sorted_pairs[0][1]}, {sorted_pairs[0][2]})..")
                for process_key in process_sum.keys():
                    pair1, pair2, fixed_col_values = [], [], []
                    values = process_sum[process_key]
                    for val in values:
                        pair1.append(val[0])
                        pair2.append(val[1])
                    # find the relevant fixed_col_values for the values
                    for s_pair in sorted_pairs:
                        df = pairs_dict[s_pair]
                        if process_key in df.columns:
                            fixed_col_values.append(s_pair[3])

                    fig, ax = plt.subplots()
                    ax.plot(fixed_col_values, pair1, marker='o', label=pair[0])
                    ax.plot(fixed_col_values, pair2, marker='o', label=pair[1])
                    for i, txt in enumerate(range(len(fixed_col_values))):
                        plt.text(fixed_col_values[i], pair1[i] + 0.04, str(round(pair1[i], 4)), fontsize=8)
                        plt.text(fixed_col_values[i], pair2[i] + 0.04, str(round(pair2[i], 4)), fontsize=8)

                    ax.set_xlabel(fixed_col)
                    ax.set_ylabel('values')
                    ax.legend()
                    title = f'{pair[0]} {pair[1]} - PROCESS {process_key}'
                    ax.set_title(title)

                    data_min = min(min(pair1), min(pair2))
                    data_max = max(max(pair1), max(pair2))
                    plt.ylim(data_min - 0.5, data_max + 0.5)

                    save_directory = os.path.join(cell_path, 'Graphs', sorted_pairs[0][1] + ", " + sorted_pairs[0][2])
                    os.makedirs(save_directory, exist_ok=True)
                    save_path = os.path.join(save_directory, f' Process {process_key} - by {fixed_col}.png')
                    plt.savefig(save_path, dpi=300)
                    plt.close(fig)

                    print(
                        f"graph {sorted_pairs[0][1]}, {sorted_pairs[0][2]}, process {process_key} saved successfully")

                create_graphs(process_sum, sorted_pairs, pairs_dict, pair, fixed_col, cell_path, cell_name)

    if not graphs:
        print("There is not enough data for creating graphs")


def plot_G_values(title: str, uid: list, values: list, save_path: str, edge_percents: float):
    """
    This function accepts columns representing processes and sorts for each process its proteins.
    In addition, the function saves the plot of process

    :param title: The plot title.
    :param uid: The sorted list of G_UID.
    :param values: The sorted list of G_values.
    :param save_path: The path to save the figures, if None the plots will be displayed one by one
    :param edge_percents: The percentage of proteins to be considered as the edge for each process.
    """
    valid.is_valid_path(save_path)

    plt.title(title)
    plt.figure(figsize=(50, 30))
    plt.scatter(uid, values)
    plt.xticks(uid, [f'{name} ({i})' for i, name in enumerate(uid)], rotation=90, fontsize=6)
    plt.ylabel('Effect')

    for i, txt in enumerate(range(len(uid))):
        plt.text(uid[i], values[i] + 0.002, str(i), fontsize=5)

    lower_edge, lower_value, upper_edge, upper_value = hf.find_edges(uid, values, edge_percents)
    if lower_edge != [] and upper_edge != []:
        lower_edge_indices = [uid.index(val) for val in lower_edge]
        upper_edge_indices = [uid.index(val) for val in upper_edge]
        plt.scatter([uid[i] for i in lower_edge_indices], [values[i] for i in lower_edge_indices], color='red')
        plt.scatter([uid[i] for i in upper_edge_indices], [values[i] for i in upper_edge_indices], color='red')

    plt.savefig(os.path.join(save_path, f'{title}.SVG'), dpi=300)
    print(f"The SVG file '{title}' saved successfully")
    plt.close()


def create_new_sheet(df: pd.DataFrame, path: str, sheet_name: str):
    """
    This function inserts the DataFrame into a new sheet in an existing Excel file.

    :param df: The DataFrame to insert into a new sheet.
    :param path: The file path.
    :param sheet_name: The name of the new sheet.
    """
    if not df.empty:
        try:
            wb = openpyxl.load_workbook(path)
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                wb.remove(sheet)
                wb.save(path)
        except Exception as e:
            print(f"Error occurred while removing the sheet: {e}")
            return
        try:
            with pd.ExcelWriter(path, mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name)
                workbook = writer.book
                worksheet = workbook[sheet_name]
                worksheet.freeze_panes = "A2"
                set_column_width_and_alignment(worksheet)

            print(f"The sheet '{sheet_name}' created successfully")
        except Exception as e:
            print(f"Error occurred while creating the sheet: {e}")
    else:
        print(f"The sheet '{sheet_name}' was not created because the DataFrame is empty")


def get_folder_name(data_path: str) -> str:
    """
    This function accepts a path with a filename at the end and returns the filename at the end of the path.
    :param data_path: The file path with the name of the folder at the end.
    :return: Name of the folder.
    """
    folder_name = os.path.basename(data_path)
    folder_name, _ = os.path.splitext(folder_name)
    return folder_name


def export_data(file_iter: int, pairs_df: pd.DataFrame, pairs_dict: dict, cell_name: str, fixed_col: str,
                data_path: str, save_path: str, sheet_name: str):
    """
    Export the analyzed data and create plots for a specific cell line.

    :param file_iter: The iteration number of the file being processed.
    :param pairs_df: The DataFrame containing the analyzed pairs.
    :param pairs_dict: The dictionary of Pandas dataframes containing the analyzed pairs.
    :param cell_name: The name of the cell line being analyzed.
    :param fixed_col: The name of the column that remains fixed in each pair.
    :param data_path: The path where the original dataframes are stored.
    :param save_path: The path where the exported data and plots will be saved.
    :param sheet_name: The name of the sheet or file to be exported.
    """
    folder_name = get_folder_name(data_path)
    folder_path = os.path.join(save_path, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    cell_path = os.path.join(folder_path, cell_name)
    os.makedirs(cell_path, exist_ok=True)

    if file_iter == 0:
        create_plots(pairs_dict, cell_name, fixed_col, cell_path)

    file_path = os.path.join(cell_path, sheet_name + '.csv')
    print(f"creating '{sheet_name}.csv'..")

    pairs_df.to_csv(file_path, index=True)
    print(f"{sheet_name}.csv created successfully\n")
